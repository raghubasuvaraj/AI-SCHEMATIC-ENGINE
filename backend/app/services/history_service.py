"""
Service for storing and managing query history.
Stores questions, generated plans, SQL queries, and results for bulk export.
"""
import io
import json
from datetime import datetime
from typing import List, Dict, Any, Optional, Literal
from pydantic import BaseModel
from backend.app.storage.state_store import JsonState

history_store = JsonState("backend/app/storage/query_history.json")


class QueryHistoryItem(BaseModel):
    """Single query history entry."""
    id: str
    timestamp: str
    database: str
    question: str
    intent: Optional[str] = None
    plan: Optional[Dict[str, Any]] = None
    sql: Optional[str] = None
    params: Optional[List] = None
    row_count: Optional[int] = None
    execution_time_ms: Optional[int] = None
    success: bool = True
    error: Optional[str] = None
    status: Literal["success", "failed", "partial"] = "success"


class QueryHistory(BaseModel):
    """Collection of query history items."""
    items: List[QueryHistoryItem] = []


def get_history() -> QueryHistory:
    """Get all query history."""
    data = history_store.load(default={"items": []})
    return QueryHistory.model_validate(data)


def save_history(history: QueryHistory) -> None:
    """Save query history."""
    history_store.save(history.model_dump())


def add_to_history(item: QueryHistoryItem) -> QueryHistory:
    """Add a new item to history."""
    history = get_history()
    # Keep last 100 items
    history.items = [item] + history.items[:99]
    save_history(history)
    return history


def clear_history() -> QueryHistory:
    """Clear all history."""
    history = QueryHistory(items=[])
    save_history(history)
    return history


def export_history_to_json() -> str:
    """Export all history as JSON string."""
    history = get_history()
    return json.dumps(history.model_dump(), indent=2)


def export_history_to_csv() -> str:
    """Export all history as CSV string."""
    history = get_history()
    if not history.items:
        return "id,timestamp,database,question,intent,query_plan,sql,status,execution_time_ms,row_count\n"
    
    lines = ["id,timestamp,database,question,intent,query_plan,sql,status,execution_time_ms,row_count"]
    for item in history.items:
        # Escape quotes in strings
        question = item.question.replace('"', '""')
        sql = (item.sql or "").replace('"', '""').replace('\n', ' ')
        plan_str = json.dumps(item.plan).replace('"', '""') if item.plan else ""
        status = item.status if hasattr(item, 'status') else ("success" if item.success else "failed")
        lines.append(f'"{item.id}","{item.timestamp}","{item.database}","{question}","{item.intent or ""}","{plan_str}","{sql}","{status}",{item.execution_time_ms or 0},{item.row_count or 0}')
    
    return "\n".join(lines)


def export_history_to_excel() -> bytes:
    """Export all history as Excel bytes (xlsx format)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # If openpyxl not installed, return CSV as fallback
        return export_history_to_csv().encode('utf-8')
    
    history = get_history()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Query History"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["#", "Timestamp", "Question", "Intent", "Query Plan", "SQL", "Status", "Time (ms)", "Rows"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, item in enumerate(history.items, 2):
        status = item.status if hasattr(item, 'status') else ("success" if item.success else "failed")
        plan_json = json.dumps(item.plan, indent=2) if item.plan else ""
        
        row_data = [
            row_idx - 1,
            item.timestamp[:19] if item.timestamp else "",  # Trim microseconds
            item.question,
            item.intent or "N/A",
            plan_json,  # Full plan JSON
            (item.sql or "")[:500],  # Truncate long SQL
            status.upper(),
            item.execution_time_ms or 0,
            item.row_count or 0
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Status coloring
            if col == 7:  # Status column
                if value == "SUCCESS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == "FAILED":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 80  # Wider for full plan JSON
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 8
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_history_table() -> List[Dict[str, Any]]:
    """Get history formatted as a table (list of dicts) for UI display."""
    history = get_history()
    table = []
    
    for idx, item in enumerate(history.items, 1):
        status = item.status if hasattr(item, 'status') else ("success" if item.success else "failed")
        
        table.append({
            "row_num": idx,
            "id": item.id,
            "timestamp": item.timestamp,
            "question": item.question,
            "intent": item.intent or "N/A",
            "query_plan": item.plan,  # Return full plan JSON
            "sql": item.sql,
            "status": status,
            "execution_time_ms": item.execution_time_ms or 0,
            "row_count": item.row_count or 0,
            "error": item.error
        })
    
    return table


def export_history_to_text() -> str:
    """Export all history as formatted text."""
    history = get_history()
    output = ["# Query History Export", f"Generated: {datetime.utcnow().isoformat()}", "=" * 60, ""]
    
    for i, item in enumerate(history.items, 1):
        output.append(f"## Query {i}")
        output.append(f"**ID:** {item.id}")
        output.append(f"**Timestamp:** {item.timestamp}")
        output.append(f"**Database:** {item.database}")
        output.append(f"**Question:** {item.question}")
        output.append(f"**Intent:** {item.intent or 'N/A'}")
        output.append(f"**Rows:** {item.row_count or 0}")
        output.append(f"**Time:** {item.execution_time_ms or 0}ms")
        output.append(f"**Success:** {item.success}")
        if item.sql:
            output.append(f"\n**SQL:**\n```sql\n{item.sql}\n```")
        if item.plan:
            output.append(f"\n**Plan:**\n```json\n{json.dumps(item.plan, indent=2)}\n```")
        if item.error:
            output.append(f"\n**Error:** {item.error}")
        output.append("\n" + "-" * 60 + "\n")
    
    return "\n".join(output)

